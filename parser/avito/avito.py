import requests
from openpyxl import Workbook
from bs4 import BeautifulSoup as BS


class AvitoParser:
    def __init__(self):
        self.session = requests.Session()
        self.session.headers = {
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.150 Safari/537.36'
        }
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws['A1'] = 'Название'
        self.ws['B1'] = 'Цена'
        self.ws['C1'] = 'url'

    def get_page(self, page: int = None):
        params = {
            'user': 1,
            'radius': 0,
            'q': 'kugo',
            'cd': 1
        }
        if page and page > 1:
            params['page'] = page
        # url = 'https://www.avito.ru/ekaterinburg/avtomobili'
        url = 'https://www.avito.ru/ekaterinburg/tovary_dlya_detey_i_igrushki/kupit-velosipedy_i_samokaty-ASgBAgICAUT~AaCGAQ'
        response = self.session.get(url, params=params)
        return response.text

    def parse_block(self, item):
        try:
            title = item.select_one('h3').string.strip()
            price = item.select_one('span.price-text-1HrJ_.text-text-1PdBw.text-size-s-1PUdo').text
            host = 'https://www.avito.ru'
            url = item.select_one(
                'a.link-link-39EVK.link-design-default-2sPEv.title-root-395AQ.iva-item-title-1Rmmj.title-list-1IIB_.title-root_maxHeight-3obWc').get(
                'href')
            city = item.select_one('div.geo-georeferences-3or5Q.text-text-1PdBw.text-size-s-1PUdo > span').string
            print(f'{city} - {title} - {price} - {host + url}')
            self.write_excel(title, price, host+url, city)
        except AttributeError:
            print(f'ОШИБКА {AttributeError}')

    def write_excel(self, title, price, url, city):
        price = price.replace(' ₽', '')
        self.ws.insert_rows(2)
        self.ws['A2'] = title
        self.ws['B2'] = price
        self.ws['C2'] = url
        # self.ws['D2'] = city
        self.wb.save('avito.xlsx')

    def get_pagination(self):
        page = self.get_page()
        soup = BS(page, 'lxml')
        limit = soup.find_all('span', class_='pagination-item-1WyVp')[-2].string
        if limit:
            return int(limit)
        else:
            return 1

    def get_block(self, page: int = None):
        page = self.get_page(page=page)
        soup = BS(page, 'lxml')
        # div = soup.select('div.items-items-38oUm')
        div = soup.find_all(attrs={'data-marker': 'item'})
        for item in div:
            self.parse_block(item)

    def parse_all(self):
        limit = self.get_pagination()
        print(f'Страниц - {limit}')
        for i in range(1, limit + 1):
            self.get_block(page=i)


def main():
    parser = AvitoParser()
    parser.parse_all()

    # parser.get_pagination()
    # parser.get_block()


if __name__ == '__main__':
    main()
