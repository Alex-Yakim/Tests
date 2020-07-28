import requests
from bs4 import BeautifulSoup as BS
from openpyxl import Workbook
import os, shutil
from multiprocessing import Pool


HOST = 'https://thermex.ru'
series_urls = []
models_urls = []
name = []
price = []
article = []
download = int(input('Скачать фото? 1 - да; 2 - нет ->'))
if download == 1: #Создание каталога для скачивания фото
    try:
        shutil.rmtree('photo-thermex')
    except:
        pass
    os.mkdir('photo-thermex')


def get_series_urls():
    r = requests.get(HOST + '/catalog')
    soup = BS(r.content, 'lxml')
    #item = soup.select('div.catalog-des > ul > li > a')
    for i in soup.select('div.catalog-des > ul > li > a'):
        series_urls.append(HOST + i.get('href'))


def get_models_urls():
    for u in series_urls:
        r = requests.get(u)
        soup = BS(r.content, 'lxml')
        for i in soup.select('div.goods-img'):
            models_urls.append(HOST + i.a['href'])


def parse_model():
    for u in models_urls:
        print(f'Парсинг товаров {models_urls.index(u) + 1} из {len(models_urls)} {u}')
        r = requests.get(u)
        soup = BS(r.content, 'lxml')
        try:
            name.append(soup.select('h1.catalog-good-title')[0].string)
        except:
            name.append('no data')
        try:
            price.append(soup.select('div.catalog-good-price.relative > big')[0].string)
        except:
            price.append('no data')
        try:
            article.append(soup.select('body > main > section.anchor-content.catalog-good-props > div > div.row > div:nth-child(1) > div:nth-child(1) > span:nth-child(2)')[0].string)
        except:
            article.append('no data')

        if u.startswith('https://thermex.ru/catalog/seriya-') and download == 1:
            download_img(soup)


def download_img(soup):
    img_url = []
    global direct
    direct = soup.select('h1.catalog-good-title')[0].string
    os.mkdir('photo-thermex/' + str(direct))
    img_dom = soup.select('div.catalog-good-gallery-root > a')
    for i in img_dom:
        img_url.append(HOST + i['href'])
    for url in img_url:
        p = requests.get(url)
        with open('photo-thermex/' + str(direct) + '/' + str(img_url.index(url)) + '.jpg', 'wb') as out:
            out.write(p.content)


def write_excel():
    print('Запись данных')
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Артикул'
    ws['B1'] = 'Название'
    ws['C1'] = 'Цена'
    for i in range(len(name)):
        ws['A' + str(i+2)] = article[i]
        ws['C' + str(i+2)] = price[i]
        ws['B' + str(i+2)] = name[i]
    wb.save('thermex_parser.xlsx')


def main():
    print('Начало парсинга')
    get_series_urls()
    print('Поиск товаров')
    get_models_urls()
    print(f'Товаров - {len(models_urls)}')
    parse_model()
    print(f'articles - {len(article)}. prices - {len(price)}. names - {len(name)}')
    print('Парсинг завершен')
    #write_excel()
    #os.startfile(r'thermex_parser.xlsx')
    input()


if __name__ == '__main__':
    main()
