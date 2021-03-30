from main import ParserInRed

import requests
from bs4 import BeautifulSoup as BS
from openpyxl import Workbook
import threading


class ParserCatalog():

    def __init__(self):
        self.session = requests.Session()
        self.session.headers = {
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.150 Safari/537.36'
        }
        self.HOST = 'https://inredhome.ru'
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws['A1'] = 'Название'
        self.ws['B1'] = 'Цена'
        self.ws['C1'] = 'Производитель'
        self.ws['D1'] = 'Каталог'
        self.ws['E1'] = 'Категория'
        self.ws['F1'] = 'Ссылка'

    def get_catalog_links(self, url):
        """Собирает ссылки на разделы каталога"""
        r = self.session.get(url)
        soup = BS(r.content, 'lxml')
        div_item = soup.select('div.section_list > div.item > div > div.name > a')
        for item in div_item:
            href = f"{self.HOST}{item.get('href')}"
            for i in range(1, self.get_pagination(href) + 1):
                self.parse_item(href, i)

    def get_pagination(self, url):
        """Возвращает количество страниц с товаром"""
        r = self.session.get(url)
        soup = BS(r.content, 'lxml')
        pagination = soup.select('#pagination > li')
        if pagination:
            max_page = pagination[-1].a.text
            return int(max_page)
        else:
            return 1

    def parse_item(self, url, page: int):
        """Парсинг товара"""
        params = {'PAGEN_4': page}
        r = self.session.get(url, params=params)
        print(r.url)
        soup = BS(r.content, 'lxml')
        div_item = soup.select('div.products__container.js_ct_product_container > div.products__item')
        for item in div_item:
            link = f"{self.HOST}{item.find('a').get('href')}"
            catalog = link.replace(self.HOST, '').split('/')[2]
            category = link.replace(self.HOST, '').split('/')[-3]
            name = item.find('div', class_='products__name').text.strip()
            price = item.find('div', class_='products__base_price').text.strip().replace('₽', '')
            price = int(price.replace(' ', ''))
            try:
                old_price = item.find('span', class_='products__old_price').text.strip().replace('Р', '')
                old_price = int(old_price.replace(' ', ''))
            except:
                old_price = None
            brand = item.find('div', class_='products__prop_item').text
            brand = brand.replace('Производитель:', '').strip()
            data = {
                'name': name,
                'url': link,
                'price': price,
                'old_price': old_price,
                'brand': brand,
                'catalog': catalog,
                'category': category,
            }
            self.write_excel(data=data)

    def write_excel(self, data: dict):
        """Запись данных"""
        self.ws.insert_rows(2)
        self.ws['A2'] = data['name']
        self.ws['B2'] = data['price']
        self.ws['C2'] = data['brand']
        self.ws['D2'] = data['catalog']
        self.ws['E2'] = data['category']
        self.ws['F2'] = data['url']
        self.wb.save('catalog.xlsx')


if __name__ == '__main__':
    cat = ParserCatalog()
    cat.get_catalog_links('https://inredhome.ru/catalog/')