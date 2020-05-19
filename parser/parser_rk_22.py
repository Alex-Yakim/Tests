import requests
from bs4 import BeautifulSoup as BS
from openpyxl import Workbook
import os, shutil

start_page = input('Начальная страница каталога -> ')
download = int(input('Скачать фото? 1 - да; 2 - нет ->'))
HOST = 'https://www.rusklimat.ru'
article = []
price = []
name = []
urls = []
status = []
item_url = []
span_status = None

if download == 1: #Создание каталога для скачивания фото
    try:
        shutil.rmtree('photo')
    except:
        pass
    os.mkdir('photo')

'''#получение значения статуса "под заказ" (не актуально)
def span_v():
    global span_status
    r = requests.get('https://www.rusklimat.ru/ekaterinburg/uvlazhniteli-i-ochistiteli-vozdukha/uvlazhniteli-vozdukha/page-3/')
    soup = BS(r.content, 'lxml')
    span = soup.select('#catalog_items > .item.b-line > div > div.c > div.cln-bsk > div.cln-bsk-b > div.block-status > span:nth-child(2)')
    span_status = span[0]'''

#расчет количества страниц
def pages_count(html):
    r = requests.get(html, params=None)
    page_bs = BS(r.content, 'lxml')
    pagination = page_bs.select('div.paginator > ul > li > a')
    if pagination:
        return int(pagination[-1].string)
    else:
        return 1



#запрос всех страниц в каталоге
def get_urls():
    print('Начало парсинга')
    for x in range(1, pages_count(start_page)+1):
        urls.append(requests.get(start_page + 'page-' + str(x)))
    else:
        print('Urls - ' + str(len(urls)))

#поиск всех товаров
def get_item_url():
    for u in urls:
        page_bs = BS(u.content, 'lxml')
        for i in page_bs.select('#catalog_items > .item.b-line > div.w > div.ttl > a'):
            item_url.append(HOST + i.get('href'))
    print('Товаров - ' + str(len(item_url)))

#парсинг товаров
def item_parse():
    for i in item_url:
        print(f'Парсинг товара {item_url.index(i) + 1} из {len(item_url)}')
        r = requests.get(i)
        soup = BS(r.content, 'lxml')
        name.append(soup.select('h1.ttl.element__header')[0].string) #Получение названия
        article.append(soup.select('div.article > span.textspan > b')[0].string) #Получение артикула
        price.append(soup.select('div.prices > div.price')[0].string.strip()) #Получение цены
        st = soup.find_all(attrs={'ajax-status': 'PREORDER'}) #Получение статуса
        if len(st) == 0:
            status.append('в наличии')
        else:
            status.append('под заказ')
        if download == 1:
            download_img(soup)
            save_params(soup)

#скачивание картинок
def download_img(soup):
    img_url = []
    global direct
    direct = soup.select('div.article > span.textspan > b')[0].string
    #os.mkdir('D:/Python/test1/parser/photo/' + str(direct))
    os.mkdir('photo/' + str(direct))
    img_dom = soup.select('#cardGal-Nav > li > a')
    for i in img_dom:
        img_url.append(HOST + i.get('href'))
    for d in img_url:
        if d.startswith('https://www.rusklimat.ruhttps://www.youtube.com'):
            img_url.remove(d)
    for url in img_url:
        p = requests.get(url)
        #with open('D:/Python/test1/parser/photo/' + str(direct) + '/' + str(img_url.index(url)) + '.jpg', 'wb') as out:
        with open('photo/' + str(direct) + '/' + str(img_url.index(url)) + '.jpg', 'wb') as out:
            out.write(p.content)

#Запись характеристик и описания
def save_params(soup):
    params = soup.select('#tabChar > div.tab-hide > table > tr > td')
    discription = soup.select('#tabDesc > div.tab-hide > article')
    with open('photo/' + str(direct) + '/' + 'характеристики.txt', 'w', encoding='utf-8') as par_out:
        for p in params:
            par_out.write(p.text + '\n')
    with open('photo/' + str(direct) + '/' + 'описание.txt', 'w', encoding='utf-8') as dis_out:
        for d in discription:
            dis_out.write(d.text + '\n')

#Запись в файл
def excel_write():
    print('Запись данных')
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Артикул'
    ws['B1'] = 'Название'
    ws['C1'] = 'Цена'
    ws['D1'] = 'Наличие'
    for i in range(len(article)):
        ws['A' + str(i+2)] = article[i]
        ws['C' + str(i+2)] = price[i]
        ws['B' + str(i+2)] = name[i]
        ws['D' + str(i+2)] = status[i]
    wb.save('parser_rk.xlsx')

def main():
    get_urls()
    get_item_url()
    item_parse()
    excel_write()
    print('Парсинг завершен')
    os.startfile(r'parser_rk.xlsx')
    input()



if __name__ == '__main__':
    main()
