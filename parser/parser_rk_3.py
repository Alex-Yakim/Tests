import requests
from bs4 import BeautifulSoup as BS
from openpyxl import Workbook
import os, shutil
import threading


start_page = input('Начальная страница каталога -> ')
download = int(input('Скачать фото? 1 - да; 2 - нет ->'))
HOST = 'https://www.rusklimat.ru'
urls = []
item_url = []
soups = []

#Создание книги excel
wb = Workbook()
ws = wb.active
ws['A1'] = 'Артикул'
ws['B1'] = 'Название'
ws['C1'] = 'Цена'
ws['D1'] = 'Наличие'

#Создание каталога для скачивания фото
if download == 1:
    try:
        shutil.rmtree('photo')
    except:
        pass
    os.mkdir('photo')


#расчет количества страниц
def pages_count(html):
    r = requests.get(html, params=None)
    page_bs = BS(r.content, 'lxml')
    pagination = page_bs.select('div.paginator > ul > li > a')
    if pagination:
        return int(pagination[-1].string)
    else:
        return 1


#запрос всех страниц в каталоге
def get_urls():
    print('Начало парсинга')
    for x in range(1, pages_count(start_page)+1):
        urls.append(requests.get(start_page + 'page-' + str(x)))
    else:
        print('Urls - ' + str(len(urls)))


#поиск всех товаров
def get_item_url():
    for u in urls:
        page_bs = BS(u.content, 'lxml')
        for i in page_bs.select('#catalog_items > .item.b-line > div.w > div.ttl > a'):
            item_url.append(HOST + i.get('href'))
    print('Товаров - ' + str(len(item_url)))


#Парсинг страницы товара
def item_parse(url, iter):
    print(f'Парсинг товара {iter + 1} из {len(item_url)}')
    r = requests.get(url)
    soup = BS(r.content, 'lxml')
    soups.append(soup) #Добавляем для последующего скачивания картинок и характеристик
    name = soup.select('h1.ttl.element__header')[0].string #Получение названия
    article = soup.select('div.article > span.textspan > b')[0].string #Получение артикула
    price = soup.select('div.prices > div.price')[0].string.strip() #Получение цены
    st = soup.find_all(attrs={'ajax-status': 'PREORDER'}) #Получение статуса
    if len(st) == 0:
        status = 'в наличии'
    else:
        status = 'под заказ'
    excel_write(article, price, name, status, iter)


#Запись в открытую книгу excel
def excel_write(article, price, name, status, i):
    ws['A' + str(i+2)] = article
    ws['C' + str(i+2)] = price
    ws['B' + str(i+2)] = name
    ws['D' + str(i+2)] = status


#скачивание картинок
def download_img(soup):
    img_url = []
    global direct
    direct = soup.select('div.article > span.textspan > b')[0].string
    os.mkdir('photo/' + str(direct))
    img_dom = soup.select('#cardGal-Nav > li > a')
    for i in img_dom:
        img_url.append(HOST + i.get('href'))
    for d in img_url:
        if d.startswith('https://www.rusklimat.ruhttps://www.youtube.com'):
            img_url.remove(d)
    for url in img_url:
        p = requests.get(url)
        with open('photo/' + str(direct) + '/' + str(img_url.index(url)) + '.jpg', 'wb') as out:
            out.write(p.content)

#Запись характеристик и описания
def save_params(soup):
    params = soup.select('#tabChar > div.tab-hide > table > tr > td')
    discription = soup.select('#tabDesc > div.tab-hide > article')
    with open('photo/' + str(direct) + '/' + 'характеристики.txt', 'w', encoding='utf-8') as par_out:
        for p in params:
            par_out.write(p.text + '\n')
    with open('photo/' + str(direct) + '/' + 'описание.txt', 'w', encoding='utf-8') as dis_out:
        for d in discription:
            dis_out.write(d.text + '\n')


def main():
    get_urls()
    get_item_url()
    #Создание потоков для парсинга
    thread_list = []
    for i in range(len(item_url)):
        t = threading.Thread(target=item_parse, args=(item_url[i], i))
        thread_list.append(t)
        t.start()
    for t in thread_list:
        t.join()
    #Скачивание картинок и характеристик
    if download == 1:
        for s in soups:
            download_img(s)
            save_params(s)
    #Сохраняем и открываем книгу excel с данными
    wb.save('parser_rk.xlsx')
    os.startfile(r'parser_rk.xlsx')

if __name__ == '__main__':
    main()
    print('Парсинг завершен')
    input()
