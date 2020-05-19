import requests
from bs4 import BeautifulSoup as BS
from openpyxl import Workbook
import os, shutil

start_page = input('Начальная страница каталога -> ')
download = int(input('Скачать фото? 1 - да; 2 - нет ->'))
HOST = 'https://shop.geizer.com'
urls = []
item_url = []
name = []
article = []
price = []

if download == 1: #Создание каталога для скачивания фото
    try:
        shutil.rmtree('photo_gz')
    except:
        pass
    os.mkdir('photo_gz')

#расчет количества страниц
def pages_count(html):
    r = requests.get(html, params=None)
    page_bs = BS(r.content, 'lxml')
    pagination = page_bs.select('#container > div.content-holder > div > section > div:nth-child(2) > div.n-pagin__list > a')
    if pagination:
        return int(pagination[-1].string)
    else:
        return 1

#запрос всех страниц в каталоге
def get_urls():
    print('Начало парсинга')
    for x in range(1, pages_count(start_page)+1):
        urls.append(requests.get(start_page, params={'PAGEN_2': str(x)}))
    else:
        print('Urls - ' + str(len(urls)))

#поиск всех товаров
def get_item_url():
    for u in urls:
        page_bs = BS(u.content, 'lxml')
        for i in page_bs.select('a.n-cat__card-img'):
            item_url.append(HOST + i.get('href'))
    print('Товаров - ' + str(len(item_url)))

#парсинг товаров
def item_parse():
    for i in item_url:
        print(f'Парсинг товара {item_url.index(i) + 1} из {len(item_url)}')
        r = requests.get(i)
        soup = BS(r.content, 'lxml')
        name.append(soup.select('#content > article > h1')[0].string) #Получение названия
        article.append(soup.select('#content > article > div.product-box > div > div.opis > div.product_article')[0].text.replace('Артикул: ', '').strip()) #Получение артикула
        if len(soup.select('.product_offers > dl.cost > dd')) == 0: #получение цены
            price.append('-')
        else:
            price.append(soup.select('.product_offers > dl.cost > dd')[0].string)
        if download == 1:
            download_img(soup)
            save_params(soup)

#скачивание картинок
def download_img(soup):
    img_url = []
    global direct
    direct = soup.select('#content > article > div.product-box > div > div.opis > div.product_article')[0].text.replace('Артикул: ', '').strip()
    os.mkdir('photo_gz/' + str(direct))
    img_dom = soup.select('div.photo > div.slide > a')
    for i in img_dom:
        img_url.append(HOST + i.get('href'))
    for url in img_url:
        p = requests.get(url)
        with open('photo_gz/' + str(direct) + '/' + str(img_url.index(url)) + '.jpg', 'wb') as out:
            out.write(p.content)

#Запись характеристик и описания
def save_params(soup):
    params = soup.select('#tab2')
    discription = soup.select('#tab1 > div.open-box')
    with open('photo_gz/' + str(direct) + '/' + 'характеристики.txt', 'w', encoding='utf-8') as par_out:
        for p in params:
            par_out.write(p.text + '\n')
    with open('photo_gz/' + str(direct) + '/' + 'описание.txt', 'w', encoding='utf-8') as dis_out:
        for d in discription:
            dis_out.write(d.text + '\n')

#Запись в файл
def excel_write():
    print('Запись данных')
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Артикул'
    ws['B1'] = 'Название'
    ws['C1'] = 'Цена'
    for i in range(len(article)):
        ws['A' + str(i+2)] = article[i]
        ws['C' + str(i+2)] = price[i]
        ws['B' + str(i+2)] = name[i]
    wb.save('parser_gz.xlsx')

def main():
    get_urls()
    get_item_url()
    item_parse()
    excel_write()
    print('Парсинг завершен')
    os.startfile(r'parser_gz.xlsx')
    input()

if __name__ == '__main__':
    main()
